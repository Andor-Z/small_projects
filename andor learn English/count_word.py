import re
import openpyxl


def getwords(filename):
    words_counts = {}
    fr = open(filename,encoding='utf-8')
    lines = fr.readlines()
    for line in lines:
        line = line.strip()
        words=re.compile(r'[^A-Z^a-z]+').split(line)
        words = [word.lower() for word in words if word!='']
        for word in words:
            words_counts[word] = words_counts.get(word, 0) + 1
    sortedwords = sorted(words_counts.items(), key = lambda w:w[1], reverse = True)
    # for word in sortedwords:
    #     print(word)
    print('一共有' + len(sortedwords) + '个单词！')
    return sortedwords

def get_mywords():
    word_set = set()
    with open('mywords.txt', 'rt', encoding='utf-8') as f:
        words = f.readlines()       
        for word in words:
            word_set.add(word)
        print(word_set)
        print('读出已完成')
    return word_set

def add2mywords(excelwt):
    word_set = get_mywords()
    book = openpyxl.load_workbook(excelwt)
    sheet = book.active
    rows = len(sheet.rows) 
    for i in range(1, rows):
        if sheet.cell(row=i, column=3).value == 5:
            # print(str(sheet.cell(row=i, column=1)))
            word_set.add(sheet.cell(row=i, column=1).value)
    print(word_set)
    print('认识的单词已经收集！')
    with open('mywords.txt', 'wt', encoding='utf-8') as f:
        for word in word_set:
            f.write(str(word) + '\n')





def write2exl(excelwt, data):
    word_set = get_mywords()
    book = openpyxl.load_workbook(excelwt)
    sheet = book.active
    data_len = len(data)

    for i in range(data_len):
        sheet.cell(row=(i+1), column=1).value = data[i][0]
        sheet.cell(row=(i+1), column=2).value = data[i][1]
        if i in word_set:
            sheet.cell(row=(i+1), column=3).value = 5
        elif len(data[i][0]) <= 3:
            sheet.cell(row=(i+1), column=3).value = 10
        elif len(data[i][0]) == 4:
            sheet.cell(row=(i+1), column=3).value = 9
        elif int(data[i][1]) == 1:
            sheet.cell(row=(i+1), column=3).value = 8
        elif int(data[i][1]) <= 3:
            sheet.cell(row=(i+1), column=3).value = 4
        else:
            sheet.cell(row=(i+1), column=3).value = 0
    book.save(excelwt)

if __name__ == '__main__':
    # data = getwords('PythonCookbook3rd.txt')
    excelwt = 'PythonCookbook3rd.xlsx'
    # write2exl(excelwt, data)
    add2mywords(excelwt)

  
