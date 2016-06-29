import xlrd
import openpyxl
import os




def find_all_file(file_dir):
	os.chdir(file_dir)
	files =os.listdir(file_dir)
	f1 = files.copy()
	if 'all_files.txt' not in files:
		f = open('all_files.txt','w')
		f.write('all_files.txt\n')
		f.write('2016微信订单明细.xlsx\n')
		for i in files:
			exc_n = i + '\n'
			f.write(exc_n)
	else:
		f = open('all_files.txt','r+')
		lines = f.readlines()
		for i in files:
			exc_n = i + '\n'
			if exc_n in lines:
				f1.remove(i)
			else:
				f.write(exc_n)
	f.close()
	#print(f1)
	return f1

def read_excel(excelrd):
	data = xlrd.open_workbook(excelrd)
	table = data.sheets()[0]
	nrows = table.nrows
	l = []
	for i in range(1,nrows):
		t = table.row_values(i)
		l.append(t)
	return l	
	
def write_data(excelwt, data):
	book = openpyxl.load_workbook(excelwt)
	sheet = book.active
	rows = len(sheet.rows)
	n = len(data)
	for i in range(n):
		n1 = len(data[i])
		for t in range(n1):
			sheet.cell(row = (i+1 + rows), column = (t+1)).value = data[i][t]
	book.save(excelwt)




def rest_data():
	os.chdir(f_dir)
	file_name = find_all_file(f_dir)
	data = []
	for i in file_name:
		l = read_excel(i)
		data.extend(l)
	#print(data)
	write_data(excelwt, data)



#datawt = xlwt3.Workbook()
#sheet = datawt.add_sheet('Sheet1')
#datawt.save(excelwt)



f_dir = 'C:\\workspace\\weixin\\发货明细'
os.chdir(f_dir)
excelrd = '刘凤俐1.4号发货单.XLSX'
excelwt = '2016微信订单明细.xlsx'
#read_excel(excelrd)

rest_data()







#a = find_all_file(f_dir)
#print(a)
